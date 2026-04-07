#!/usr/bin/env python3
"""
Shopify Daily Sales Reporter
Her sabah otomatik olarak tüm mağazalardan satış verilerini çeker,
PDF + Excel raporları oluşturur ve WhatsApp'tan gönderir.
"""

import json, os, sys, requests, schedule, time, logging
from datetime import datetime, timedelta
from pathlib import Path

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).parent
CONFIG_PATH = BASE_DIR / "config.json"
REPORTS_DIR = BASE_DIR / "reports"
REPORTS_DIR.mkdir(exist_ok=True)


def load_config():
    with open(CONFIG_PATH) as f:
        return json.load(f)


def get_access_token(store):
    """Client credentials grant ile access token al (24 saat geçerli)."""
    url = f"https://{store['shop_url']}/admin/oauth/access_token"
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    data = {
        "client_id": store["client_id"],
        "client_secret": store["client_secret"],
        "grant_type": "client_credentials"
    }
    resp = requests.post(url, headers=headers, data=data, timeout=30)
    resp.raise_for_status()
    return resp.json()["access_token"]


def fetch_orders(store, token, since_date, until_date=None):
    """Belirli tarih aralığındaki siparişleri çek."""
    orders = []
    url = f"https://{store['shop_url']}/admin/api/2024-10/orders.json"
    params = {
        "status": "any",
        "created_at_min": since_date.isoformat(),
        "limit": 250,
        "fields": "id,name,created_at,total_price,line_items,financial_status,currency"
    }
    if until_date:
        params["created_at_max"] = until_date.isoformat()

    while url:
        resp = requests.get(url, headers={"X-Shopify-Access-Token": token}, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        orders.extend(data.get("orders", []))
        # Pagination
        link = resp.headers.get("Link", "")
        url = None
        params = None
        if 'rel="next"' in link:
            for part in link.split(","):
                if 'rel="next"' in part:
                    url = part.split(";")[0].strip().strip("<>")
    return orders


def get_yesterday_range():
    """Dünün başlangıç ve bitiş zamanlarını döndür."""
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday = today - timedelta(days=1)
    return yesterday, today


def get_period_range(days):
    """Son X günün tarih aralığını döndür."""
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    start = today - timedelta(days=days)
    return start, today


def process_orders_summary(orders):
    """Sipariş özetini çıkar: toplam adet, toplam tutar."""
    total_items = 0
    total_amount = 0.0
    for order in orders:
        if order.get("financial_status") in ("paid", "partially_paid", "authorized", None):
            total_amount += float(order.get("total_price", 0))
            for item in order.get("line_items", []):
                total_items += item.get("quantity", 0)
    return total_items, total_amount


def process_orders_detail(orders):
    """Ürün bazlı detay çıkar."""
    products = {}
    for order in orders:
        if order.get("financial_status") in ("paid", "partially_paid", "authorized", None):
            for item in order.get("line_items", []):
                name = item.get("title", "Bilinmeyen Ürün")
                qty = item.get("quantity", 0)
                price = float(item.get("price", 0))
                if name in products:
                    products[name]["qty"] += qty
                    products[name]["total"] += price * qty
                else:
                    products[name] = {"qty": qty, "price": price, "total": price * qty}
    return products


# ==================== RAPOR 1: GÜNLÜK ÖZET ====================

def generate_report1_pdf(all_data, report_date, currency):
    """Rapor 1: Mağaza bazlı günlük satış özeti PDF."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    filepath = REPORTS_DIR / f"rapor1_gunluk_ozet_{report_date}.pdf"
    doc = SimpleDocTemplate(str(filepath), pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=18, spaceAfter=20)
    story = []

    story.append(Paragraph(f"Günlük Satış Özeti - {report_date}", title_style))
    story.append(Spacer(1, 10))

    table_data = [["Mağaza", "Satılan Ürün Adedi", f"Toplam Satış ({currency})"]]
    grand_items = 0
    grand_total = 0.0

    for store_name, data in all_data.items():
        items, amount = data["summary"]
        table_data.append([store_name, str(items), f"{currency}{amount:,.2f}"])
        grand_items += items
        grand_total += amount

    table_data.append(["GENEL TOPLAM", str(grand_items), f"{currency}{grand_total:,.2f}"])

    table = Table(table_data, colWidths=[7*cm, 4.5*cm, 5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#27AE60')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F2F3F4')]),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(table)
    doc.build(story)
    return filepath


def generate_report1_excel(all_data, report_date, currency):
    """Rapor 1: Mağaza bazlı günlük satış özeti Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    filepath = REPORTS_DIR / f"rapor1_gunluk_ozet_{report_date}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Günlük Özet"

    header_fill = PatternFill('solid', fgColor='2C3E50')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    total_fill = PatternFill('solid', fgColor='27AE60')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:C1')
    ws['A1'] = f"Günlük Satış Özeti - {report_date}"
    ws['A1'].font = Font(bold=True, size=16)

    headers = ["Mağaza", "Satılan Ürün Adedi", f"Toplam Satış ({currency})"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    row = 4
    for store_name, data in all_data.items():
        items, amount = data["summary"]
        ws.cell(row=row, column=1, value=store_name).border = border
        ws.cell(row=row, column=2, value=items).border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        c = ws.cell(row=row, column=3, value=amount)
        c.number_format = f'"{currency}"#,##0.00'
        c.border = border
        c.alignment = Alignment(horizontal='center')
        row += 1

    for col in range(1, 4):
        cell = ws.cell(row=row, column=col)
        cell.fill = total_fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=1, value="GENEL TOPLAM")
    ws.cell(row=row, column=2).value = f"=SUM(B4:B{row-1})"
    ws.cell(row=row, column=3).value = f"=SUM(C4:C{row-1})"
    ws.cell(row=row, column=3).number_format = f'"{currency}"#,##0.00'

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22

    wb.save(str(filepath))
    return filepath


# ==================== RAPOR 2: ÜRÜN DETAY ====================

def generate_report2_pdf(all_data, report_date, currency):
    """Rapor 2: Mağaza bazlı ürün detay raporu PDF."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    filepath = REPORTS_DIR / f"rapor2_urun_detay_{report_date}.pdf"
    doc = SimpleDocTemplate(str(filepath), pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=18, spaceAfter=20)
    store_style = ParagraphStyle('StoreTitle', parent=styles['Heading2'], fontSize=14, spaceAfter=10,
                                  textColor=colors.HexColor('#2C3E50'))
    story = []

    story.append(Paragraph(f"Ürün Detay Raporu - {report_date}", title_style))
    story.append(Spacer(1, 10))

    first = True
    for store_name, data in all_data.items():
        if not first:
            story.append(Spacer(1, 20))
        first = False

        story.append(Paragraph(f"📦 {store_name}", store_style))
        products = data["detail"]

        if not products:
            story.append(Paragraph("Bu tarihte satış yapılmamıştır.", styles['Normal']))
            continue

        table_data = [["Ürün Adı", "Adet", f"Birim Fiyat ({currency})", f"Toplam ({currency})"]]
        store_total_items = 0
        store_total_amount = 0.0

        for pname, pinfo in sorted(products.items()):
            table_data.append([
                Paragraph(pname, styles['Normal']),
                str(pinfo["qty"]),
                f"{currency}{pinfo['price']:,.2f}",
                f"{currency}{pinfo['total']:,.2f}"
            ])
            store_total_items += pinfo["qty"]
            store_total_amount += pinfo["total"]

        table_data.append([
            "TOPLAM", str(store_total_items), "",
            f"{currency}{store_total_amount:,.2f}"
        ])

        table = Table(table_data, colWidths=[7*cm, 2.5*cm, 3.5*cm, 3.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1ABC9C')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F9FA')]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(table)

    doc.build(story)
    return filepath


def generate_report2_excel(all_data, report_date, currency):
    """Rapor 2: Mağaza bazlı ürün detay raporu Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    filepath = REPORTS_DIR / f"rapor2_urun_detay_{report_date}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill('solid', fgColor='34495E')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    total_fill = PatternFill('solid', fgColor='1ABC9C')
    store_font = Font(bold=True, size=13, color='2C3E50')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for store_name, data in all_data.items():
        safe_name = store_name[:31]
        ws = wb.create_sheet(title=safe_name)

        ws.merge_cells('A1:D1')
        ws['A1'] = f"{store_name} - Ürün Detay - {report_date}"
        ws['A1'].font = store_font

        headers = ["Ürün Adı", "Adet", f"Birim Fiyat ({currency})", f"Toplam ({currency})"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        products = data["detail"]
        row = 4
        for pname, pinfo in sorted(products.items()):
            ws.cell(row=row, column=1, value=pname).border = border
            ws.cell(row=row, column=2, value=pinfo["qty"]).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            c = ws.cell(row=row, column=3, value=pinfo["price"])
            c.number_format = f'"{currency}"#,##0.00'
            c.border = border
            c = ws.cell(row=row, column=4, value=pinfo["total"])
            c.number_format = f'"{currency}"#,##0.00'
            c.border = border
            row += 1

        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = total_fill
            cell.font = Font(bold=True, color='FFFFFF')
            cell.border = border
        ws.cell(row=row, column=1, value="TOPLAM")
        ws.cell(row=row, column=2).value = f"=SUM(B4:B{row-1})"
        ws.cell(row=row, column=4).value = f"=SUM(D4:D{row-1})"
        ws.cell(row=row, column=4).number_format = f'"{currency}"#,##0.00'

        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

    wb.save(str(filepath))
    return filepath


# ==================== RAPOR 3: DÖNEMSEL ====================

def generate_report3_pdf(all_period_data, currency):
    """Rapor 3: Son 1/3/6 ay satış raporu PDF."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    today_str = datetime.now().strftime("%Y-%m-%d")
    filepath = REPORTS_DIR / f"rapor3_donemsel_{today_str}.pdf"
    doc = SimpleDocTemplate(str(filepath), pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=18, spaceAfter=15)
    period_style = ParagraphStyle('PeriodTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=10,
                                   textColor=colors.HexColor('#8E44AD'))
    store_style = ParagraphStyle('StoreTitle', parent=styles['Heading2'], fontSize=13, spaceAfter=8,
                                  textColor=colors.HexColor('#2C3E50'))
    story = []

    story.append(Paragraph(f"Dönemsel Satış Raporu - {today_str}", title_style))

    period_labels = {"30": "Son 1 Ay", "90": "Son 3 Ay", "180": "Son 6 Ay"}

    for period_key, period_label in period_labels.items():
        story.append(Spacer(1, 15))
        story.append(Paragraph(f"📅 {period_label}", period_style))

        period_data = all_period_data.get(period_key, {})
        store_totals = []

        for store_name, data in period_data.items():
            story.append(Paragraph(f"  {store_name}", store_style))
            products = data["detail"]

            if not products:
                story.append(Paragraph("    Bu dönemde satış yok.", styles['Normal']))
                store_totals.append((store_name, 0, 0.0))
                continue

            table_data = [["Ürün Adı", "Adet", f"Birim Fiyat ({currency})", f"Toplam ({currency})"]]
            s_items = 0
            s_amount = 0.0
            for pname, pinfo in sorted(products.items(), key=lambda x: -x[1]["total"]):
                table_data.append([
                    Paragraph(pname, styles['Normal']),
                    str(pinfo["qty"]),
                    f"{currency}{pinfo['price']:,.2f}",
                    f"{currency}{pinfo['total']:,.2f}"
                ])
                s_items += pinfo["qty"]
                s_amount += pinfo["total"]

            table_data.append(["TOPLAM", str(s_items), "", f"{currency}{s_amount:,.2f}"])
            store_totals.append((store_name, s_items, s_amount))

            table = Table(table_data, colWidths=[7*cm, 2.5*cm, 3.5*cm, 3.5*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8E44AD')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1ABC9C')),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F9FA')]),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(table)
            story.append(Spacer(1, 10))

        # Dönem toplam tablosu
        story.append(Spacer(1, 10))
        story.append(Paragraph(f"📊 {period_label} - Mağaza Toplamları", store_style))
        sum_table = [["Mağaza", "Toplam Adet", f"Toplam Satış ({currency})"]]
        g_items = 0
        g_amount = 0.0
        for sn, si, sa in store_totals:
            sum_table.append([sn, str(si), f"{currency}{sa:,.2f}"])
            g_items += si
            g_amount += sa
        sum_table.append(["GENEL TOPLAM", str(g_items), f"{currency}{g_amount:,.2f}"])

        t = Table(sum_table, colWidths=[7*cm, 4*cm, 5.5*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E74C3C')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ]))
        story.append(t)
        story.append(PageBreak())

    doc.build(story)
    return filepath


def generate_report3_excel(all_period_data, currency):
    """Rapor 3: Son 1/3/6 ay satış raporu Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    today_str = datetime.now().strftime("%Y-%m-%d")
    filepath = REPORTS_DIR / f"rapor3_donemsel_{today_str}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill('solid', fgColor='8E44AD')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    total_fill = PatternFill('solid', fgColor='1ABC9C')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    period_labels = {"30": "Son 1 Ay", "90": "Son 3 Ay", "180": "Son 6 Ay"}

    for period_key, period_label in period_labels.items():
        ws = wb.create_sheet(title=period_label)
        period_data = all_period_data.get(period_key, {})

        ws.merge_cells('A1:D1')
        ws['A1'] = f"{period_label} Satış Raporu - {today_str}"
        ws['A1'].font = Font(bold=True, size=14, color='8E44AD')

        row = 3
        for store_name, data in period_data.items():
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row=row, column=1, value=store_name).font = Font(bold=True, size=12, color='2C3E50')
            row += 1

            headers = ["Ürün Adı", "Adet", f"Birim Fiyat ({currency})", f"Toplam ({currency})"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            row += 1

            products = data["detail"]
            start_row = row
            for pname, pinfo in sorted(products.items(), key=lambda x: -x[1]["total"]):
                ws.cell(row=row, column=1, value=pname).border = border
                ws.cell(row=row, column=2, value=pinfo["qty"]).border = border
                ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
                c = ws.cell(row=row, column=3, value=pinfo["price"])
                c.number_format = f'"{currency}"#,##0.00'
                c.border = border
                c = ws.cell(row=row, column=4, value=pinfo["total"])
                c.number_format = f'"{currency}"#,##0.00'
                c.border = border
                row += 1

            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.fill = total_fill
                cell.font = Font(bold=True, color='FFFFFF')
                cell.border = border
            ws.cell(row=row, column=1, value="MAĞAZA TOPLAMI")
            if row > start_row:
                ws.cell(row=row, column=2).value = f"=SUM(B{start_row}:B{row-1})"
                ws.cell(row=row, column=4).value = f"=SUM(D{start_row}:D{row-1})"
                ws.cell(row=row, column=4).number_format = f'"{currency}"#,##0.00'
            row += 2

        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

    wb.save(str(filepath))
    return filepath


# ==================== MESAJ GÖNDERME ====================

def send_whatsapp(config, message, files=None):
    """Twilio ile WhatsApp mesajı gönder. Uzun mesajları böler."""
    wa = config.get("whatsapp", {})
    sid = wa.get("twilio_sid")
    token = wa.get("twilio_token")
    from_num = wa.get("from_number")
    to_num = wa.get("to_number")

    if not all([sid, token, from_num, to_num]) or "TWILIO" in sid:
        return False

    try:
        from twilio.rest import Client
        client = Client(sid, token)
        
        # Mesaj 1500 karakterden uzunsa böl
        if len(message) <= 1500:
            client.messages.create(body=message, from_=from_num, to=to_num)
        else:
            parts = []
            current = ""
            for line in message.split("\n"):
                if len(current) + len(line) + 1 > 1400:
                    parts.append(current)
                    current = line
                else:
                    current += "\n" + line if current else line
            if current:
                parts.append(current)
            
            for i, part in enumerate(parts):
                if len(parts) > 1:
                    part = f"({i+1}/{len(parts)})\n{part}"
                client.messages.create(body=part, from_=from_num, to=to_num)
        
        logger.info("✅ WhatsApp mesajı gönderildi")
        if files:
            for f in files:
                logger.info(f"Rapor kaydedildi: {f}")
        return True
    except Exception as e:
        logger.error(f"WhatsApp gönderilemedi: {e}")
        return False


def send_email(config, subject, message, files=None):
    """Gmail ile e-posta gönder (dosya ekleriyle)."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    email_cfg = config.get("email", {})
    sender = email_cfg.get("gmail_address", "")
    app_password = email_cfg.get("gmail_app_password", "")
    recipients = email_cfg.get("to_emails", [])

    if not sender or not app_password or sender == "SENIN_GMAIL_ADRESIN@gmail.com":
        return False

    msg = MIMEMultipart()
    msg['From'] = sender
    msg['To'] = ", ".join(recipients)
    msg['Subject'] = subject
    msg.attach(MIMEText(message, 'plain', 'utf-8'))

    if files:
        for filepath in files:
            filepath = Path(filepath)
            if filepath.exists():
                with open(filepath, 'rb') as f:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename="{filepath.name}"')
                msg.attach(part)

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(sender, app_password)
            server.sendmail(sender, recipients, msg.as_string())
        logger.info(f"✅ E-posta gönderildi: {', '.join(recipients)}")
        return True
    except Exception as e:
        logger.error(f"E-posta gönderilemedi: {e}")
        return False


def send_report(config, subject, message, files=None):
    """WhatsApp ve/veya Email ile gönder (hangisi yapılandırılmışsa)."""
    sent = False
    sent = send_whatsapp(config, message, files) or sent
    sent = send_email(config, subject, message, files) or sent
    if not sent:
        logger.warning("Ne WhatsApp ne Email yapılandırılmış. Raporlar sadece yerel olarak kaydedildi.")


# ==================== ANA FONKSİYON ====================

def run_daily_report():
    """Günlük raporu çalıştır."""
    logger.info("=== Günlük Satış Raporu Başlıyor ===")
    config = load_config()
    stores = config["stores"]
    currency = config.get("currency_symbol", "£")
    yesterday_start, yesterday_end = get_yesterday_range()
    report_date = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")

    all_data = {}

    for store in stores:
        name = store["name"]
        logger.info(f"  → {name} verisi çekiliyor...")
        try:
            token = get_access_token(store)
            orders = fetch_orders(store, token, yesterday_start, yesterday_end)
            summary = process_orders_summary(orders)
            detail = process_orders_detail(orders)
            all_data[name] = {"summary": summary, "detail": detail, "orders": orders}
            logger.info(f"    ✓ {name}: {summary[0]} ürün, {currency}{summary[1]:,.2f}")
        except Exception as e:
            logger.error(f"    ✗ {name} hatası: {e}")
            all_data[name] = {"summary": (0, 0.0), "detail": {}, "orders": []}

    # Rapor 1
    logger.info("Rapor 1 oluşturuluyor (Günlük Özet)...")
    r1_pdf = generate_report1_pdf(all_data, report_date, currency)
    r1_xlsx = generate_report1_excel(all_data, report_date, currency)

    # Rapor 2
    logger.info("Rapor 2 oluşturuluyor (Ürün Detay)...")
    r2_pdf = generate_report2_pdf(all_data, report_date, currency)
    r2_xlsx = generate_report2_excel(all_data, report_date, currency)

    # WhatsApp mesajı
    msg_lines = [f"📊 *Günlük Satış Raporu* - {report_date}\n"]
    grand_items = 0
    grand_total = 0.0
    for sn, sd in all_data.items():
        items, amount = sd["summary"]
        msg_lines.append(f"🏪 *{sn}*: {items} ürün → {currency}{amount:,.2f}")
        grand_items += items
        grand_total += amount
    msg_lines.append(f"\n💰 *GENEL TOPLAM*: {grand_items} ürün → {currency}{grand_total:,.2f}")

    # Son 30 günlük özet ekle
    logger.info("Son 30 günlük veri çekiliyor...")
    month_start, month_end = get_period_range(30)
    msg_lines.append(f"\n\n📅 *Son 30 Gün Satış Özeti*\n")
    m_grand_items = 0
    m_grand_total = 0.0
    for store in stores:
        name = store["name"]
        try:
            token = get_access_token(store)
            m_orders = fetch_orders(store, token, month_start, month_end)
            m_items, m_amount = process_orders_summary(m_orders)
            msg_lines.append(f"🏪 *{name}*: {m_items} ürün → {currency}{m_amount:,.2f}")
            m_grand_items += m_items
            m_grand_total += m_amount
        except Exception as e:
            logger.error(f"    ✗ {name} 30 gün hatası: {e}")
            msg_lines.append(f"🏪 *{name}*: Veri alınamadı")
    msg_lines.append(f"\n💰 *30 GÜN TOPLAM*: {m_grand_items} ürün → {currency}{m_grand_total:,.2f}")

    send_report(config, f"Günlük Satış Raporu - {report_date}", "\n".join(msg_lines), [r1_pdf, r1_xlsx, r2_pdf, r2_xlsx])
    logger.info(f"✅ Raporlar kaydedildi: {REPORTS_DIR}")
    return [r1_pdf, r1_xlsx, r2_pdf, r2_xlsx]


def run_periodic_report():
    """Dönemsel raporu çalıştır (Rapor 3)."""
    logger.info("=== Dönemsel Satış Raporu Başlıyor ===")
    config = load_config()
    stores = config["stores"]
    currency = config.get("currency_symbol", "£")

    all_period_data = {}
    for days_key, days in [("30", 30), ("90", 90), ("180", 180)]:
        start, end = get_period_range(days)
        period_data = {}
        for store in stores:
            name = store["name"]
            logger.info(f"  → {name} ({days_key} gün) verisi çekiliyor...")
            try:
                token = get_access_token(store)
                orders = fetch_orders(store, token, start, end)
                summary = process_orders_summary(orders)
                detail = process_orders_detail(orders)
                period_data[name] = {"summary": summary, "detail": detail}
                logger.info(f"    ✓ {name}: {summary[0]} ürün")
            except Exception as e:
                logger.error(f"    ✗ {name} hatası: {e}")
                period_data[name] = {"summary": (0, 0.0), "detail": {}}
        all_period_data[days_key] = period_data

    r3_pdf = generate_report3_pdf(all_period_data, currency)
    r3_xlsx = generate_report3_excel(all_period_data, currency)

    send_report(config, f"Dönemsel Satış Raporu - {datetime.now().strftime('%Y-%m-%d')}", "Dönemsel Satış Raporu ekte.", [r3_pdf, r3_xlsx])
    logger.info(f"✅ Dönemsel rapor kaydedildi: {REPORTS_DIR}")
    return [r3_pdf, r3_xlsx]


# ==================== RAPOR 4: BİRLEŞİK ÜRÜN LİSTESİ ====================

def generate_combined_pdf(combined_products, report_date, currency):
    """Tüm mağazalardan satılan ürünler - mağaza adıyla birlikte - PDF."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    filepath = REPORTS_DIR / f"rapor4_birlesik_urunler_{report_date}.pdf"
    doc = SimpleDocTemplate(str(filepath), pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=18, spaceAfter=20)
    story = []

    if "30gun" in report_date:
        title_text = f"Son 30 Gün Satılan Ürünler - {report_date.replace('_30gun','')}"
    else:
        title_text = f"Günlük Satılan Ürünler - {report_date}"
    story.append(Paragraph(title_text, title_style))
    story.append(Spacer(1, 15))

    # Özet tablosu ekle (mağaza bazlı toplamlar)
    if combined_products:
        summary_style = ParagraphStyle('SummaryTitle', parent=styles['Heading2'], fontSize=14, spaceAfter=10)
        story.append(Paragraph("📊 Mağaza Özeti", summary_style))
        store_summary = {}
        for item in combined_products:
            s = item["store"]
            if s not in store_summary:
                store_summary[s] = {"qty": 0, "total": 0.0}
            store_summary[s]["qty"] += item["qty"]
            store_summary[s]["total"] += item["total"]
        
        sum_data = [["Mağaza", "Toplam Adet", f"Toplam Satış ({currency})"]]
        sg_items = 0
        sg_total = 0.0
        for sname, sdata in store_summary.items():
            sum_data.append([sname, str(sdata["qty"]), f"{currency}{sdata['total']:,.2f}"])
            sg_items += sdata["qty"]
            sg_total += sdata["total"]
        sum_data.append(["GENEL TOPLAM", str(sg_items), f"{currency}{sg_total:,.2f}"])
        
        sum_table = Table(sum_data, colWidths=[6*cm, 4*cm, 6*cm])
        sum_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#27AE60')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F2F3F4')]),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(sum_table)
        story.append(Spacer(1, 25))
        story.append(Paragraph("📦 Ürün Detayları", summary_style))
        story.append(Spacer(1, 10))

    table_data = [["Mağaza", "Ürün", "Adet", f"Fiyat ({currency})", f"Toplam ({currency})"]]
    grand_items = 0
    grand_total = 0.0

    for item in sorted(combined_products, key=lambda x: (x["store"], -x["total"])):
        short_name = Paragraph(item["product"], styles['Normal'])
        table_data.append([
            Paragraph(item["store"], styles['Normal']),
            short_name,
            str(item["qty"]),
            f"{currency}{item['price']:,.2f}",
            f"{currency}{item['total']:,.2f}"
        ])
        grand_items += item["qty"]
        grand_total += item["total"]

    table_data.append(["", "TOPLAM", str(grand_items), "", f"{currency}{grand_total:,.2f}"])

    table = Table(table_data, colWidths=[3*cm, 7*cm, 1.5*cm, 2.5*cm, 2.8*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A5276')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E74C3C')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#EBF5FB')]),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(table)
    doc.build(story)
    return filepath


def generate_combined_excel(combined_products, report_date, currency):
    """Tüm mağazalardan satılan ürünler - mağaza adıyla birlikte - Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    filepath = REPORTS_DIR / f"rapor4_birlesik_urunler_{report_date}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Birleşik Ürün Listesi"

    header_fill = PatternFill('solid', fgColor='1A5276')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    total_fill = PatternFill('solid', fgColor='E74C3C')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:E1')
    if "30gun" in report_date:
        ws['A1'] = f"Son 30 Gün Satılan Ürünler - {report_date.replace('_30gun','')}"
    else:
        ws['A1'] = f"Günlük Satılan Ürünler - {report_date}"
    ws['A1'].font = Font(bold=True, size=14, color='1A5276')

    headers = ["Mağaza", "Ürün", "Adet", f"Birim Fiyat ({currency})", f"Toplam ({currency})"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    row = 4
    for item in sorted(combined_products, key=lambda x: (x["store"], -x["total"])):
        c = ws.cell(row=row, column=1, value=item["store"])
        c.border = border
        c.alignment = Alignment(wrap_text=True)
        c = ws.cell(row=row, column=2, value=item["product"])
        c.border = border
        c.alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=3, value=item["qty"]).border = border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        c = ws.cell(row=row, column=4, value=item["price"])
        c.number_format = f'"{currency}"#,##0.00'
        c.border = border
        c = ws.cell(row=row, column=5, value=item["total"])
        c.number_format = f'"{currency}"#,##0.00'
        c.border = border
        row += 1

    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.fill = total_fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=2, value="TOPLAM")
    ws.cell(row=row, column=3).value = f"=SUM(C4:C{row-1})"
    ws.cell(row=row, column=5).value = f"=SUM(E4:E{row-1})"
    ws.cell(row=row, column=5).number_format = f'"{currency}"#,##0.00'

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16

    wb.save(str(filepath))
    return filepath


def run_combined_report():
    """Tüm mağazalardan günlük + son 30 günlük satılan ürünleri listele."""
    logger.info("=== Birleşik Ürün Raporu Başlıyor ===")
    config = load_config()
    stores = config["stores"]
    currency = config.get("currency_symbol", "£")
    yesterday_start, yesterday_end = get_yesterday_range()
    report_date = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    month_start, month_end = get_period_range(30)

    daily_products = []
    monthly_products = []

    for store in stores:
        name = store["name"]
        logger.info(f"  → {name} günlük verisi çekiliyor...")
        try:
            token = get_access_token(store)
            # Günlük veri
            orders = fetch_orders(store, token, yesterday_start, yesterday_end)
            store_products = {}
            for order in orders:
                for item in order.get("line_items", []):
                    pname = item.get("title", "Bilinmeyen Ürün")
                    qty = item.get("quantity", 0)
                    price = float(item.get("price", 0))
                    if pname in store_products:
                        store_products[pname]["qty"] += qty
                        store_products[pname]["total"] += price * qty
                    else:
                        store_products[pname] = {"qty": qty, "price": price, "total": price * qty}
            for pname, pinfo in store_products.items():
                daily_products.append({
                    "store": name, "product": pname,
                    "qty": pinfo["qty"], "price": pinfo["price"], "total": pinfo["total"]
                })

            # 30 günlük veri
            logger.info(f"  → {name} aylık verisi çekiliyor...")
            m_orders = fetch_orders(store, token, month_start, month_end)
            m_store_products = {}
            for order in m_orders:
                for item in order.get("line_items", []):
                    pname = item.get("title", "Bilinmeyen Ürün")
                    qty = item.get("quantity", 0)
                    price = float(item.get("price", 0))
                    if pname in m_store_products:
                        m_store_products[pname]["qty"] += qty
                        m_store_products[pname]["total"] += price * qty
                    else:
                        m_store_products[pname] = {"qty": qty, "price": price, "total": price * qty}
            for pname, pinfo in m_store_products.items():
                monthly_products.append({
                    "store": name, "product": pname,
                    "qty": pinfo["qty"], "price": pinfo["price"], "total": pinfo["total"]
                })

            logger.info(f"    ✓ {name} tamamlandı")
        except Exception as e:
            logger.error(f"    ✗ {name} hatası: {e}")

    # Günlük PDF + Excel
    r4_pdf = generate_combined_pdf(daily_products, report_date, currency)
    r4_xlsx = generate_combined_excel(daily_products, report_date, currency)

    # 30 günlük PDF + Excel
    r5_pdf = generate_combined_pdf(monthly_products, f"{report_date}_30gun", currency)
    r5_xlsx = generate_combined_excel(monthly_products, f"{report_date}_30gun", currency)

    grand_items = sum(p["qty"] for p in daily_products)
    grand_total = sum(p["total"] for p in daily_products)
    m_grand_items = sum(p["qty"] for p in monthly_products)
    m_grand_total = sum(p["total"] for p in monthly_products)

    msg = f"📦 *Birleşik Ürün Raporu* - {report_date}\n\n"
    
    store_totals = {}
    for item in daily_products:
        s = item["store"]
        if s not in store_totals:
            store_totals[s] = {"qty": 0, "total": 0.0}
        store_totals[s]["qty"] += item["qty"]
        store_totals[s]["total"] += item["total"]
    
    for sname, sdata in store_totals.items():
        msg += f"🏪 *{sname}*: {sdata['qty']} ürün → {currency}{sdata['total']:,.2f}\n"
    
    msg += f"\n💰 *GÜNLÜK TOPLAM: {grand_items} ürün → {currency}{grand_total:,.2f}*\n"
    msg += f"📅 *30 GÜN TOPLAM: {m_grand_items} ürün → {currency}{m_grand_total:,.2f}*\n"
    msg += f"\n📄 Detaylı ürün listesi PDF raporlarında."

    send_report(config, f"Birleşik Ürün Raporu - {report_date}", msg, [r4_pdf, r4_xlsx, r5_pdf, r5_xlsx])
    logger.info(f"✅ Birleşik rapor kaydedildi: {REPORTS_DIR}")
    return [r4_pdf, r4_xlsx, r5_pdf, r5_xlsx]


def main():
    if len(sys.argv) > 1:
        cmd = sys.argv[1]
        if cmd == "daily":
            run_daily_report()
        elif cmd == "periodic":
            run_periodic_report()
        elif cmd == "combined":
            run_combined_report()
        elif cmd == "all":
            run_daily_report()
            run_combined_report()
            run_periodic_report()
        elif cmd == "schedule":
            config = load_config()
            report_time = config.get("report_time", "07:45")
            logger.info(f"Zamanlayıcı başlatıldı. Her gün saat {report_time}'de çalışacak.")
            schedule.every().day.at(report_time).do(run_daily_report)
            schedule.every().monday.at("08:00").do(run_periodic_report)
            while True:
                schedule.run_pending()
                time.sleep(30)
        else:
            print("Kullanım: python reporter.py [daily|combined|periodic|all|schedule]")
    else:
        print("Kullanım:")
        print("  python reporter.py daily     → Günlük rapor (Rapor 1 + 2)")
        print("  python reporter.py combined  → Birleşik ürün listesi (Rapor 4)")
        print("  python reporter.py periodic  → Dönemsel rapor (Rapor 3)")
        print("  python reporter.py all       → Tüm raporlar")
        print("  python reporter.py schedule  → Otomatik zamanlayıcı")


if __name__ == "__main__":
    main()
